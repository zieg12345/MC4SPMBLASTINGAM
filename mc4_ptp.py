import streamlit as st
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import pytz

def mc4_ptp_section():
    st.subheader("MC4 PTP")
    
    uploaded_file = st.file_uploader(
        "📤 Choose an Excel file",
        type=["xlsx"],
        key="mc4_ptp_uploader",
        help="Upload an Excel (.xlsx) file with 'Client Name' and 'Email' columns"
    )
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success("File uploaded successfully!")
    
    # Reset button
    if st.session_state.uploaded_file is not None:
        if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
            st.session_state.uploaded_file = None
            st.rerun()

    if st.session_state.uploaded_file is not None:
        try:
            df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl', dtype=str)
            df = df.fillna("")  # Replace NaN with empty string
            df.columns = df.columns.str.strip()
            
            # Define required columns with "chname" instead of "Name" and "{{agentcode}}" instead of "Collector"
            required_columns = ["Email", "{{chname}}", "Account No.", "{{agentcode}}", "Financing/Card No.", "Client Name", "Product Type"]
            
            # Check for required 'Client Name' column
            if "Client Name" not in df.columns:
                st.error("Missing required column 'Client Name' in the uploaded file.")
            else:
                # Select only specified columns and fill missing with empty string
                # Map "Name" to "{{chname}}" and "Collector" to "{{agentcode}}" if they exist in the input
                if "Name" in df.columns:
                    df = df.rename(columns={"Name": "{{chname}}"})
                if "Collector" in df.columns:
                    df = df.rename(columns={"Collector": "{{agentcode}}"})
                df = df[required_columns].copy()
                
                # Filter rows where 'Email' contains '@' if 'Email' column exists
                if "Email" in df.columns:
                    df = df[df["Email"].str.contains("@", na=False)]
                
                # Group by 'Client Name' and create multi-sheet Excel
                output = BytesIO()
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet
                
                # Create ML REFERENCE sheet with all data
                ws_ml_ref = wb.create_sheet(title="ML REFERENCE")
                headers = list(df.columns)
                for col_num, header in enumerate(headers, 1):
                    ws_ml_ref.cell(row=1, column=col_num).value = header
                for row_num, row in enumerate(df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws_ml_ref.cell(row=row_num, column=col_num).value = str(value)
                        ws_ml_ref.cell(row=row_num, column=col_num).number_format = '@'
                
                # List to store unique Client Names for display
                unique_clients = []
                
                for client_name, group_df in df.groupby("Client Name"):
                    sheet_name = client_name.strip() if client_name.strip() else "Blank Client Name"
                    unique_clients.append(sheet_name)
                    if sheet_name in ["SBC CARDS RECOV L1", "SBC CARDS & LOAN L6", "SBC PL RECOV L1"]:
                        # Combine these into a single "SBC Combined" sheet
                        if "SBC Combined" not in wb.sheetnames:
                            ws_combined = wb.create_sheet(title="SBC Combined")
                            headers = list(group_df.columns)
                            for col_num, header in enumerate(headers, 1):
                                ws_combined.cell(row=1, column=col_num).value = header
                        # Append data to SBC Combined sheet with {{agentcode}} set to "FRBALBIRAN"
                        for row_num, row in enumerate(group_df.values, ws_combined.max_row + 1):
                            new_row = list(row)
                            new_row[headers.index("{{agentcode}}")] = "FRBALBIRAN"
                            for col_num, value in enumerate(new_row, 1):
                                ws_combined.cell(row=row_num, column=col_num).value = str(value)
                                ws_combined.cell(row=row_num, column=col_num).number_format = '@'
                    else:
                        # Create individual sheets for other Client Names with original {{agentcode}} values
                        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit is 31 characters
                        headers = list(group_df.columns)
                        for col_num, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_num).value = header
                        for row_num, row in enumerate(group_df.values, 2):
                            for col_num, value in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num).value = str(value)
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                
                # Display preview list (sample) for each sheet
                st.write("### Preview of Processed Data")
                for client_name in unique_clients:
                    if client_name == "SBC Combined":
                        sheet_df = df[df["Client Name"].isin(["SBC CARDS RECOV L1", "SBC CARDS & LOAN L6", "SBC PL RECOV L1"])].copy()
                        sheet_df["{{agentcode}}"] = "FRBALBIRAN"
                    else:
                        sheet_df = df[df["Client Name"] == client_name]
                    st.write(f"**{client_name}**")
                    st.dataframe(sheet_df.head(), use_container_width=True)
                
                wb.save(output)
                output.seek(0)
                
                today = datetime.now(pytz.timezone('Asia/Manila')).strftime("%B %d %Y")
                file_name = f"MC4_PTP_{today}.xlsx"
                st.download_button(
                    label="📥 Download Multi-Sheet Excel",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_mc4_ptp",
                    use_container_width=True
                )
        except Exception as e:
            st.error(f"An error occurred while processing the file: {str(e)}")
    else:
        st.info("Please upload an Excel file with 'Client Name' and 'Email' columns to divide data into multiple sheets.")