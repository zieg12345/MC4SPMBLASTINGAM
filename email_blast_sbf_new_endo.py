import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def email_blast_sbf_new_endo_section():
    st.subheader("SBF NEW ENDO Email Blast Uploader")

    uploaded_file = st.file_uploader(
        "📤 Choose SBF NEW ENDO Excel file",
        type=["xlsx"],
        key="sbf_new_endo_uploader",
        help="Upload an Excel Workbook (.xlsx) with columns: Account No., Name, Email, Collector, Financing/Card No."
    )

    if uploaded_file is not None:
        st.session_state.uploaded_sbf_new_endo_file = uploaded_file
        st.success("File uploaded successfully!")

    if st.session_state.get('uploaded_sbf_new_endo_file') is not None:
        if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
            st.session_state.uploaded_sbf_new_endo_file = None
            st.rerun()

    current_date = datetime.now().strftime("SBF NEW ENDO %b %d %Y %I_%M %p PST").upper()

    if st.session_state.get('uploaded_sbf_new_endo_file') is not None:
        try:
            df = pd.read_excel(st.session_state.uploaded_sbf_new_endo_file, engine='openpyxl', dtype=str)
            df = df.fillna("")
            df.columns = df.columns.str.strip()

            required_columns = ["Account No.", "Name", "Email", "Collector", "Financing/Card No."]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error(f"The following required columns are missing: {', '.join(missing_columns)}")
                return

            initial_row_count = len(df)
            df = df[df["Email"].str.contains("@", na=False)]
            filtered_row_count = len(df)

            if filtered_row_count < initial_row_count:
                st.info(f"Removed {initial_row_count - filtered_row_count} rows with invalid or missing email addresses.")

            summary_df = pd.DataFrame({
                "{{email}}": df["Email"],
                "{{chname}}": df["Name"],
                "{{agentcode}}": df["Collector"],
                "{{ID}}": "SCCO",
                "Account No.": df["Account No."],
                "Financing/Card No.": df["Financing/Card No."]
            })

            st.subheader("Summary Table")
            st.dataframe(summary_df, use_container_width=True)

            st.info(f"Total Rows: {len(summary_df)}")
            st.info(f"Unique Emails: {len(summary_df['{{email}}'].unique())}")
            st.info(f"Unique Names: {len(summary_df['{{chname}}'].unique())}")
            st.info(f"Unique Accounts: {len(df['Account No.'].unique())}")
            st.info(f"Unique Financing/Card Nos.: {len(df['Financing/Card No.'].unique())}")

            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "SBF NEW ENDO"
            headers = list(summary_df.columns)
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            for row_num, row in enumerate(summary_df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num).value = str(value)
                    ws.cell(row=row_num, column=col_num).number_format = '@'
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="📥 Download Summary Table as Excel",
                data=output,
                file_name=f"{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_sbf_new_endo_summary",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"An error occurred while processing the file: {str(e)}")
    else:
        st.info("Please upload an Excel file to generate the summary table.")