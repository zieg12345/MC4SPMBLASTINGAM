import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def email_blast_bucket4_section():
    st.subheader("Bucket 4 Generic Template Email Blast File Uploader")
    uploaded_file = st.file_uploader(
        "📤 Choose an Excel file",
        type=["xlsx"],
        key="bucket4_uploader",
        help="Upload an Excel (.xlsx) file with columns: Email, Name, Collector, Product Type, Financing/Card No., Account No., Assign Date"
    )
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success("File uploaded successfully!")
    if st.session_state.uploaded_file is not None:
        if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
            st.session_state.uploaded_file = None
            st.rerun()
    # Sample data for Bucket 4 Generic Template
    sample_data = {
        "Email": ["JDBenbinuto@securitybank.com.ph"],
        "{{chname}}": ["Janica d Benbinuto"],
        "{{agentcode}}": ["PJHA"],
        "{{product}}": ["CARD"],
        "Financing/Card No.": ["123456789"],
        "Account No.": ["987654321"],
        "Assign Date": [datetime.now().strftime('%Y-%m-%d')],
        "{{ID}}": ["4DCO"]
    }
    sample_df = pd.DataFrame(sample_data)
    if st.session_state.uploaded_file is not None:
        try:
            df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl', dtype=str)
            df = df.fillna("")
            # Validate Email column for '@' symbol
            initial_row_count_email = len(df)
            df = df[df['Email'].str.contains('@', na=False)]
            if len(df) < initial_row_count_email:
                st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
            required_columns = ['Email', 'Name', 'Collector', 'Product Type', 'Financing/Card No.', 'Account No.', 'Assign Date']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error(f"Missing required columns in the uploaded file: {', '.join(missing_columns)}")
            else:
                df = df[required_columns]
                column_mapping = {
                    'Name': '{{chname}}',
                    'Collector': '{{agentcode}}',
                    'Product Type': '{{product}}'
                }
                df = df.rename(columns=column_mapping)
                df['{{product}}'] = df['{{product}}'].replace({'MC': 'CARD', 'BEL': 'BUSINESS EXPRESS LOAN'})
                df['{{agentcode}}'] = df['{{agentcode}}'].apply(lambda x: 'PJHA' if x == 'SPMADRID' else x)
                df['{{ID}}'] = df['{{agentcode}}'].apply(lambda x: '4DCO' if x == 'PJHA' else '4CCO')
                # Combine with sample data
                summary_df = pd.concat([df, sample_df], ignore_index=True)
                st.write("### Processed Data")
                st.dataframe(summary_df, use_container_width=True)
                st.write("### Summary")
                total_records = len(summary_df)
                unique_emails = summary_df['Email'].nunique()
                unique_names = summary_df['{{chname}}'].nunique()
                unique_agents = summary_df['{{agentcode}}'].nunique()
                unique_products = summary_df['{{product}}'].nunique()
                unique_accounts = summary_df['Account No.'].nunique()
                unique_ids = summary_df['{{ID}}'].nunique()
                try:
                    date_range = f"From {summary_df['Assign Date'].min()} to {summary_df['Assign Date'].max()}"
                except:
                    date_range = "Invalid date format"
                st.write(f"- **Total Records**: {total_records}")
                st.write(f"- **Unique Emails**: {unique_emails}")
                st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                st.write(f"- **Unique Agents ({{agentcode}})**: {unique_agents}")
                st.write(f"- **Unique Products**: {unique_products}")
                st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                st.write(f"- **Unique IDs ({{ID}})**: {unique_ids}")
                st.write(f"- **Assign Date Range**: {date_range}")
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Summary"
                headers = list(summary_df.columns)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num).value = header
                for row_num, row in enumerate(summary_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num).value = str(value)
                        ws.cell(row=row_num, column=col_num).number_format = '@'
                wb.save(output)
                output.seek(0)
                today = datetime.now().strftime("%B %d %Y")
                file_name = f"B4 Email blasting {today}.xlsx"
                st.download_button(
                    label="📥 Download Processed Excel",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_bucket4_summary",
                    use_container_width=True
                )
        except Exception as e:
            st.error(f"An error occurred while processing the file: {str(e)}")
    else:
        st.subheader("Sample Summary Table")
        st.dataframe(sample_df, use_container_width=True)
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        headers = list(sample_df.columns)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        for row_num, row in enumerate(sample_df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num).value = str(value)
                ws.cell(row=row_num, column=col_num).number_format = '@'
        wb.save(output)
        output.seek(0)
        today = datetime.now().strftime("%B %d %Y")
        file_name = f"B4 Email blasting {today}.xlsx"
        st.download_button(
            label="📥 Download Sample Excel",
            data=output,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_bucket4_sample",
            use_container_width=True
        )
        st.info("Please upload an Excel file to generate the summary table with your data.")