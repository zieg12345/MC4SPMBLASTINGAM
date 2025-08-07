import streamlit as st
import pandas as pd
from datetime import datetime
import pytz
from openpyxl import Workbook
from io import BytesIO

@st.cache_data(show_spinner=False)
def process_excel_file(uploaded_file, stats_option):
    # Read Excel file with all columns as strings
    df = pd.read_excel(uploaded_file, engine='openpyxl', dtype=str)
    df = df.fillna("")  # Replace NaN with empty string
    df.columns = df.columns.str.strip()
    
    # Define required and optional columns
    required_columns = ["Account No.", "Name", "Financing/Card No."]
    optional_columns = ["Email"]
    missing_required = [col for col in required_columns if col not in df.columns]
    
    if missing_required:
        return None, f"Missing required columns in the uploaded file: {', '.join(missing_required)}"
    
    # Select required columns and include optional ones if present
    output_columns = required_columns + [col for col in optional_columns if col in df.columns]
    summary_df = df[output_columns].copy()
    
    # Rename columns to match output format
    column_mapping = {
        "Account No.": "ACCOUNT NUMBER",
        "Name": "NAME",
        "Financing/Card No.": "CARD NUMBER"
    }
    summary_df = summary_df.rename(columns=column_mapping)
    
    # Add required output columns
    summary_df["STATUS CODE"] = "EMAIL BLAST SENT - WAITING FOR REPLY"
    summary_df["REMARKS BY"] = "ZMJEPOLLO"
    # Apply live PHT timestamp to each row for REMARKS DATE
    summary_df["REMARKS DATE"] = summary_df.apply(
        lambda _: datetime.now(pytz.timezone('Asia/Manila')).strftime("%Y-%m-%d %H:%M:%S"), 
        axis=1
    )
    
    # Generate REMARKS column based on stats_option with fresh PHT timestamp
    if stats_option == "SBF NEGATIVE AUTOSTATS":
        summary_df["REMARKS"] = summary_df.apply(
            lambda row: f"EMAIL_SP MADRID_{datetime.now(pytz.timezone('Asia/Manila')).strftime('%Y-%m-%d %H:%M:%S')}_ZMJEPOLLO - {row['Email'] if 'Email' in row and row['Email'] else ''} NEGATIVE TEMPLATE",
            axis=1
        )
    else:  # L1-L6 NEGATIVE AUTOSTATS
        summary_df["REMARKS"] = summary_df.apply(
            lambda row: f"SPMA | 08 With SMS / email / DL without response - {row['Email'] if 'Email' in row and row['Email'] else ''} EMAIL SENT",
            axis=1
        )
    
    # Ensure all required output columns are present
    for col in ["ACCOUNT NUMBER", "NAME", "CARD NUMBER", "STATUS CODE", "REMARKS", "REMARKS BY", "REMARKS DATE"]:
        if col not in summary_df.columns:
            summary_df[col] = ""  # Add empty column if missing
    
    # Reorder columns
    summary_df = summary_df[["ACCOUNT NUMBER", "NAME", "CARD NUMBER", "STATUS CODE", "REMARKS", "REMARKS BY", "REMARKS DATE"]]
    
    # Remove duplicates based on ACCOUNT NUMBER
    initial_row_count = len(summary_df)
    summary_df = summary_df.drop_duplicates(subset=["ACCOUNT NUMBER"], keep="first")
    removed_duplicates = initial_row_count - len(summary_df)
    
    # Generate Excel output
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{stats_option.replace(' ', '_')}"
    headers = ["ACCOUNT NUMBER", "NAME", "CARD NUMBER", "STATUS CODE", "REMARKS", "REMARKS BY", "REMARKS DATE"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    for row_num, row in enumerate(summary_df.values, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num).value = str(value)
            ws.cell(row=row_num, column=col_num).number_format = '@'
    wb.save(output)
    output.seek(0)
    
    return summary_df, output, removed_duplicates

def auto_statistics_section():
    st.subheader("Auto Statistics")
    
    # Display current PHT timestamp for debugging
    current_pht = datetime.now(pytz.timezone('Asia/Manila')).strftime("%Y-%m-%d %H:%M:%S")
    st.write(f"**Current PHT Time**: {current_pht}")
    
    # Dropdown for selecting Auto Statistics option
    stats_option = st.selectbox(
        "Select Statistics Type",
        ["SBF NEGATIVE AUTOSTATS", "L1-L6 NEGATIVE AUTOSTATS"],
        help="Choose the type of statistics to generate",
        key="auto_stats_select",
        index=["SBF NEGATIVE AUTOSTATS", "L1-L6 NEGATIVE AUTOSTATS"].index(st.session_state.get("auto_stats_option", "SBF NEGATIVE AUTOSTATS"))
    )
    st.session_state.auto_stats_option = stats_option

    if stats_option in ["SBF NEGATIVE AUTOSTATS", "L1-L6 NEGATIVE AUTOSTATS"]:
        st.write(f"### {stats_option}")
        uploaded_file = st.file_uploader(
            "📤 Choose an Excel file",
            type=["xlsx"],
            key=f"{stats_option.lower().replace(' ', '_')}_uploader",
            help="Upload an Excel (.xlsx) file with columns: Account No., Name, Financing/Card No., Email (optional)"
        )
        if uploaded_file is not None:
            st.session_state.uploaded_file = uploaded_file
            st.success("File uploaded successfully!")
        
        # Reset button
        if st.session_state.uploaded_file is not None:
            if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
                st.session_state.uploaded_file = None
                st.cache_data.clear()  # Clear cache on reset
                st.rerun()

        if st.session_state.uploaded_file is not None:
            try:
                # Process the file
                summary_df, output, removed_duplicates = process_excel_file(st.session_state.uploaded_file, stats_option)
                
                if summary_df is None:
                    st.error(output)  # Output is the error message
                else:
                    if removed_duplicates > 0:
                        st.info(f"Removed {removed_duplicates} duplicate rows based on 'ACCOUNT NUMBER'.")
                    
                    st.write("### Processed Data")
                    st.dataframe(summary_df, use_container_width=True)
                    
                    # Generate summary statistics
                    st.write("### Summary")
                    total_records = len(summary_df)
                    unique_accounts = summary_df["ACCOUNT NUMBER"].nunique()
                    unique_names = summary_df["NAME"].nunique()
                    unique_cards = summary_df["CARD NUMBER"].nunique()
                    status_counts = summary_df["STATUS CODE"].value_counts().to_dict()
                    st.write(f"- **Total Records**: {total_records}")
                    st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                    st.write(f"- **Unique Names**: {unique_names}")
                    st.write(f"- **Unique Card Numbers**: {unique_cards}")
                    st.write(f"- **Status Code Distribution**: {status_counts}")
                    
                    # Download button for Excel output
                    today = datetime.now(pytz.timezone('Asia/Manila')).strftime("%B %d %Y")
                    file_name = f"{stats_option.replace(' ', '_')} {today}.xlsx"
                    st.download_button(
                        label="📥 Download Processed Excel",
                        data=output,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{stats_option.lower().replace(' ', '_')}",
                        use_container_width=True
                    )
            except Exception as e:
                st.error(f"An error occurred while processing the file: {str(e)}")
        else:
            st.info("Please upload an Excel file with columns 'Account No.', 'Name', 'Financing/Card No.', and optionally 'Email' to generate the summary table.")