import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def viber_blast_section():
    st.subheader("Viber Blast CSV Uploader")

    # File uploader for raw data (CSV)
    uploaded_file = st.file_uploader(
        "📤 Choose Raw Data CSV file",
        type=["csv"],
        key="viber_blast_uploader_raw",
        help="Upload a CSV with columns: Contact No., Debtor Name, Account No., Client, Validity"
    )
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success("Raw data file uploaded successfully!")

    # File uploader for collector lookup (Excel)
    uploaded_lookup_file = st.file_uploader(
        "📤 Choose Collector Lookup Excel file",
        type=["xlsx"],
        key="viber_blast_uploader_lookup",
        help="Upload an Excel Workbook (.xlsx) with columns: Account No., Collector"
    )
    if uploaded_lookup_file is not None:
        st.session_state.uploaded_lookup_file = uploaded_lookup_file
        st.success("Collector lookup file uploaded successfully!")

    # Reset button
    if st.session_state.get('uploaded_file') is not None or st.session_state.get('uploaded_lookup_file') is not None:
        if st.button("🔄 Reset", help="Clear the uploaded files and reset"):
            st.session_state.uploaded_file = None
            st.session_state.uploaded_lookup_file = None
            st.session_state.button1_clicked = False
            st.rerun()

    # Original sample data (used for SBC CARDS CURING B2 and default)
    original_sample_data = {
        "Campaign": ["SAMPLE", "SAMPLE", "SAMPLE", "SAMPLE"],
        "CH Code": ["12345", "123456", "1234567", "12345678"],
        "First Name": ["", "", "", ""],
        "Full Name": ["Richard Arenas", "Jinnggoy Dela Cruz", "Roman Dalisay", "Edwin Paras"],
        "Last Name": ["Collector A", "Collector B", "PJHA", "Collector D"],
        "Mobile Number": ["09274186327", "09760368821", "09088925110", "09175791122"],
        "OB": ["", "", "", ""]
    }

    # Sample data for SBC CURING B4
    sbc_curing_b4_sample_data = {
        "Campaign": ["TEST"],
        "CH Code": ["1234"],
        "First Name": [""],
        "Full Name": ["Janica d Benbinuto"],
        "Last Name": ["TEST"],
        "Mobile Number": ["09655669672"],
        "OB": [""]
    }

    # Initialize sample DataFrame with original sample data
    sample_df = pd.DataFrame(original_sample_data)

    # Dynamic filename (will be updated based on detected campaign)
    campaign_name = "GENERIC"  # Default
    current_date = datetime.now().strftime(f"VIBER BLAST {campaign_name} %b %d %Y %I_%M %p PST").upper()

    if st.session_state.get('uploaded_file') is not None:
        try:
            # Read raw data CSV
            df = pd.read_csv(st.session_state.uploaded_file, encoding='utf-8-sig', skipinitialspace=True, dtype=str)
            df = df.fillna("")  # Replace NaN with empty string
            df.columns = df.columns.str.strip()

            # Validate required columns
            required_columns = ["Contact No.", "Debtor Name", "Account No.", "Client", "Validity"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error(f"The following required columns are missing in raw data: {', '.join(missing_columns)}")
                return

            # Clean Contact No. and Account No. (remove ="" wrappers)
            df["Contact No."] = df["Contact No."].str.replace(r'^="|"$', '', regex=True)
            df["Account No."] = df["Account No."].str.replace(r'^="|"$', '', regex=True)

            # Filter out invalid rows based on Validity column
            initial_row_count_validity = len(df)
            df = df[df["Validity"] == "Valid"]
            if len(df) < initial_row_count_validity:
                st.info(f"Removed {initial_row_count_validity - len(df)} rows where Validity is not 'Valid'.")

            # Validate Contact No. length (11 digits)
            df["Contact No."] = df["Contact No."].str.strip()
            invalid_contact_no = df[df["Contact No."].str.len() != 11]
            if not invalid_contact_no.empty:
                st.warning(f"Found {len(invalid_contact_no)} rows where Contact No. is not 11 digits. These rows are still included but may need review.")

            # Filter out rows with 'BEL' in Account No.
            initial_row_count_bel = len(df)
            df = df[~df["Account No."].str.contains("BEL", case=False, na=False)]
            if initial_row_count_bel != len(df):
                st.info(f"Removed {initial_row_count_bel - len(df)} rows where Account No. contains 'BEL'.")

            # Remove duplicates based on Account No. and Contact No.
            initial_row_count = len(df)
            df = df.drop_duplicates(subset=["Account No.", "Contact No."], keep="first")
            if initial_row_count != len(df):
                st.info(f"Removed {initial_row_count - len(df)} duplicate rows based on 'Account No.' and 'Contact No.'.")

            # Prepare summary DataFrame
            summary_df = pd.DataFrame({
                "Campaign": df["Client"],
                "CH Code": df["Account No."],
                "First Name": "",
                "Full Name": df["Debtor Name"],
                "Last Name": "",  # To be populated from collector file
                "Mobile Number": df["Contact No."],
                "OB": ""
            })

            # Update sample data and filename based on Campaign
            if "SBC CURING B4" in summary_df["Campaign"].values:
                sample_df = pd.DataFrame(sbc_curing_b4_sample_data)
                campaign_name = "SBC CURING B4"
                st.info("Detected 'SBC CURING B4' in Campaign. Using custom sample data with 'Janica d Benbinuto'.")
            elif "SBC CARDS CURING B2" in summary_df["Campaign"].values:
                sample_df = pd.DataFrame(original_sample_data)
                campaign_name = "SBC CARDS CURING B2"
                st.info("Detected 'SBC CARDS CURING B2' in Campaign. Using original sample data with 'Richard Arenas' and others.")
            else:
                campaign_name = "GENERIC"
                st.info("No specific campaign detected. Using default sample data with 'Richard Arenas' and others.")

            # Update filename with detected campaign
            current_date = datetime.now().strftime(f"VIBER BLAST {campaign_name} %b %d %Y %I_%M %p PST").upper()

            # Process collector lookup file if uploaded
            if st.session_state.get('uploaded_lookup_file') is not None:
                try:
                    lookup_df = pd.read_excel(st.session_state.uploaded_lookup_file, engine='openpyxl', dtype=str)
                    lookup_df = lookup_df.fillna("")
                    lookup_df.columns = lookup_df.columns.str.strip()

                    # Validate required columns in lookup file
                    required_lookup_columns = ["Account No.", "Collector"]
                    missing_lookup_columns = [col for col in required_lookup_columns if col not in lookup_df.columns]
                    if missing_lookup_columns:
                        st.error(f"The following required columns are missing in lookup file: {', '.join(missing_lookup_columns)}")
                        return

                    # Clean Account No. in lookup file (ensure string, strip whitespace)
                    lookup_df["Account No."] = lookup_df["Account No."].str.strip()

                    # Merge with lookup file to get Collector
                    merged_df = summary_df.merge(
                        lookup_df[["Account No.", "Collector"]],
                        left_on="CH Code",
                        right_on="Account No.",
                        how="left"
                    )

                    # Update Last Name with Collector
                    summary_df["Last Name"] = merged_df["Collector"].fillna("")

                    # Report matching statistics
                    matched_count = len(merged_df[merged_df["Collector"].notna()])
                    unmatched_count = len(merged_df[merged_df["Collector"].isna()])
                    st.info(f"Collector lookup completed. {matched_count} out of {len(summary_df)} accounts matched with collectors.")
                    if unmatched_count > 0:
                        st.warning(f"{unmatched_count} accounts did not find a matching Collector in the lookup file. Check for formatting issues in 'Account No.'.")
                except Exception as e:
                    st.error(f"Error processing lookup file: {str(e)}. Please ensure the Excel file is valid and contains the required columns.")
                    return

            # Combine with sample data
            summary_df = pd.concat([summary_df, sample_df], ignore_index=True)

            # Display summary table
            st.subheader("Summary Table")
            st.dataframe(summary_df, use_container_width=True)

            # Generate Excel file
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Viber Blast"
            headers = list(summary_df.columns)
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            for row_num, row in enumerate(summary_df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num).value = str(value)
                    ws.cell(row=row_num, column=col_num).number_format = '@'
            wb.save(output)
            output.seek(0)

            # Download button
            st.download_button(
                label="📥 Download Summary Table as Excel",
                data=output,
                file_name=f"{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_summary",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"An error occurred while processing the raw data file: {str(e)}")
    else:
        # Use original sample data if no file is uploaded
        st.subheader("Sample Summary Table")
        st.dataframe(sample_df, use_container_width=True)
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Viber Blast"
        headers = list(sample_df.columns)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        for row_num, row in enumerate(sample_df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num).value = str(value)
                ws.cell(row=row_num, column=col_num).number_format = '@'
        wb.save(output)
        output.seek(0)
        st.download_button(
            label="📥 Download Sample Excel",
            data=output,
            file_name=f"{current_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_sample",
            use_container_width=True
        )
        st.info("Please upload a CSV file with raw data and an Excel file with collector data to generate the summary table.")