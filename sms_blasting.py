# sms_blasting.py
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def sms_blasting_section():
    st.subheader("SMS Blasting Uploader")

    uploaded_file = st.file_uploader(
        "Choose SMS data file (CSV or XLSX)",
        type=["csv", "xlsx"],
        key="sms_blasting_uploader",
        help="File must contain: Contact No., Account No., Debtor Name, Client"
    )

    if uploaded_file is None:
        st.info("Upload a file to start.")
        return

    # ------------------- READ FILE -------------------
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file, encoding="utf-8-sig", dtype=str)
        else:
            df = pd.read_excel(uploaded_file, engine="openpyxl", dtype=str)

        df = df.fillna("")
        df.columns = df.columns.str.strip()
    except Exception as e:
        st.error(f"Could not read the file: {e}")
        return

    # ------------------- VALIDATE COLUMNS -------------------
    required = ["Contact No.", "Account No.", "Debtor Name", "Client"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Missing required column(s): {', '.join(missing)}")
        st.info(f"Available columns: {list(df.columns)}")
        return

    # ------------------- CLEAN CONTACT NO. -------------------
    df["Contact No."] = (
        df["Contact No."]
        .astype(str)
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)          # Excel float .0
        .str.replace(r"[^0-9]", "", regex=True)        # keep only digits
    )

    # ------------------- NEW VALIDATION -------------------
    mask_starts_09 = df["Contact No."].str.startswith("09")
    mask_len_11    = df["Contact No."].str.len() >= 11
    mask_valid     = mask_starts_09 & mask_len_11

    valid_df = df[mask_valid].copy()
    removed  = len(df) - len(valid_df)

    if removed:
        st.warning(
            f"Removed {removed} row(s) that do **not** start with **09** "
            "or have fewer than 11 digits."
        )

    if valid_df.empty:
        st.error("No valid contact numbers found – nothing to export.")
        return

    # ------------------- CONVERT TO 639... -------------------
    valid_df["Mobile Number"] = (
        valid_df["Contact No."]
        .str.slice(0, 11)                # keep first 11 digits
        .str.replace(r"^09", "639", regex=True)
    )

    # ------------------- BUILD FINAL TABLE -------------------
    summary_df = pd.DataFrame({
        "Campaign":   valid_df["Client"],
        "Account No.":    valid_df["Account No."],
        "First Name": "",
        "Name":  valid_df["Debtor Name"],
        "Last Name":  "",
        "Mobile Number": valid_df["Mobile Number"],
        "OB":         ""
    })

    # ------------------- DYNAMIC FILENAME -------------------
    client_val = (
        valid_df["Client"]
        .replace("", pd.NA)
        .dropna()
        .iloc[0] if not valid_df["Client"].replace("", pd.NA).dropna().empty else "GENERIC"
    )
    campaign_name = str(client_val).strip().upper()
    file_ts = datetime.now().strftime(f"SMS BLAST {campaign_name} %b %d %Y %I_%M %p PST").upper()

    # ------------------- DISPLAY -------------------
    st.subheader("SMS Blast Summary")
    st.dataframe(summary_df, use_container_width=True)
    st.success(f"{len(summary_df)} valid record(s) ready for download.")

    # ------------------- EXCEL EXPORT -------------------
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SMS Blast"

    for c, hdr in enumerate(summary_df.columns, 1):
        ws.cell(row=1, column=c, value=hdr)

    for r, row in enumerate(summary_df.itertuples(index=False, name=None), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val))
            ws.cell(row=r, column=c).number_format = "@"

    wb.save(output)
    output.seek(0)

    st.download_button(
        label="Download SMS Blast Excel",
        data=output,
        file_name=f"{file_ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="sms_download",
        use_container_width=True
    )