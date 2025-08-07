import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def email_blast_bucket2_section():
    st.subheader("Bucket 2 with Sequence Template Email Blast File Uploader")
    uploaded_file = st.file_uploader(
        "📤 Choose a CSV or Excel file",
        type=["csv", "xlsx"],
        key="email_blast_uploader",
        help="Upload a CSV or Excel (.xlsx) file with columns: Contract Number, Email, {{chname}}, Statement Balance (OB), Statement Overdue Amount (MYP), Statement Minimum Payment (MAD), Assignment Date, TEMPLATE 1 D1, TEMPLATE 1 D2, etc."
    )
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success("Main file uploaded successfully!")
    collector_file = st.file_uploader(
        "📤 Choose a CSV or Excel file for Collector and Assign Date data",
        type=["csv", "xlsx"],
        key="collector_uploader",
        help="Upload a CSV or Excel (.xlsx) file with columns: Financing/Card No., Collector, Assign Date"
    )
    if collector_file is not None:
        st.session_state.collector_file = collector_file
        st.success("Collector file uploaded successfully!")
    if st.session_state.uploaded_file is not None or st.session_state.collector_file is not None:
        if st.button("🔄 Reset", help="Clear the uploaded files and reset"):
            st.session_state.uploaded_file = None
            st.session_state.collector_file = None
            st.session_state.button2_clicked = False
            st.rerun()
    if st.session_state.uploaded_file is not None:
        try:
            if st.session_state.uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(st.session_state.uploaded_file, encoding='utf-8-sig', skipinitialspace=True, dtype=str)
            elif st.session_state.uploaded_file.name.endswith('.xlsx'):
                df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl', dtype=str)
            df = df.fillna("")  # Replace NaN with empty string
            if df.empty:
                st.error("The uploaded file is empty. Please upload a valid file.")
                st.stop()
            df.columns = df.columns.str.strip()
            # Validate Email column for '@' symbol
            initial_row_count_email = len(df)
            df = df[df['Email'].str.contains('@', na=False)]
            if len(df) < initial_row_count_email:
                st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
            required_columns = [
                "Contract Number", "Email", "{{chname}}", "Statement Balance (OB)",
                "Statement Overdue Amount (MYP)", "Statement Minimum Payment (MAD)",
                "Assignment Date", "TEMPLATE 1 D1", "TEMPLATE 1 D2", "TEMPLATE 2 D1",
                "TEMPLATE 2 D2", "TEMPLATE 3 D1", "TEMPLATE 3 D2", "TEMPLATE 4 D1",
                "TEMPLATE 4 D2", "TEMPLATE 5 D1", "TEMPLATE 5 D2", "TEMPLATE 6 D1",
                "TEMPLATE 6 D2"
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error(f"The following required columns are missing in the main file: {', '.join(missing_columns)}")
                st.stop()
            initial_row_count = len(df)
            df = df.drop_duplicates(subset="Contract Number", keep="first")
            if initial_row_count != len(df):
                st.info(f"Removed {initial_row_count - len(df)} duplicate rows based on 'Contract Number'.")
            with st.expander("🔍 Show Detected Column Names (Main File)"):
                st.write("Detected Column Names:", list(df.columns))
            summary_df = pd.DataFrame()
            summary_df["Contract Number"] = df["Contract Number"]
            summary_df["Email"] = df["Email"]
            summary_df["{{chname}}"] = df["{{chname}}"]
            summary_df["{{agentcode}}"] = ""
            summary_df["{{ID}}"] = ""
            # Format original amount fields with commas
            def format_amount(value):
                if not value.strip():
                    return ""
                try:
                    num = float(value)
                    decimal_places = len(value.split('.')[-1]) if '.' in value else 0
                    return f"{num:,.{decimal_places}f}"
                except:
                    return value
            summary_df["{{OB}}"] = df["Statement Balance (OB)"].apply(format_amount)
            summary_df["{{MYP}}"] = df["Statement Overdue Amount (MYP)"].apply(format_amount)
            summary_df["{{MAD}}"] = df["Statement Minimum Payment (MAD)"].apply(format_amount)
            # Calculate OB+CF, MAD+CF, MYP+CF with commas and matching decimal places
            def safe_multiply(value, factor):
                if not value.strip():
                    return ""
                try:
                    num = float(value)
                    decimal_places = len(value.split('.')[-1]) if '.' in value else 0
                    result = num * factor
                    return f"{result:,.{decimal_places}f}"
                except:
                    return value
            summary_df["{{OB+CF}}"] = df["Statement Balance (OB)"].apply(lambda x: safe_multiply(x, 1.11))
            summary_df["{{MAD+CF}}"] = df["Statement Minimum Payment (MAD)"].apply(lambda x: safe_multiply(x, 1.11))
            summary_df["{{MYP+CF}}"] = df["Statement Overdue Amount (MYP)"].apply(lambda x: safe_multiply(x, 1.11))
            summary_df["TEMPLATE 1 D1"] = df["TEMPLATE 1 D1"]
            summary_df["TEMPLATE 1 D2"] = df["TEMPLATE 1 D2"]
            summary_df["TEMPLATE 2 D1"] = df["TEMPLATE 2 D1"]
            summary_df["TEMPLATE 2 D2"] = df["TEMPLATE 2 D2"]
            summary_df["TEMPLATE 3 D1"] = df["TEMPLATE 3 D1"]
            summary_df["TEMPLATE 3 D2"] = df["TEMPLATE 3 D2"]
            summary_df["TEMPLATE 4 D1"] = df["TEMPLATE 4 D1"]
            summary_df["TEMPLATE 4 D2"] = df["TEMPLATE 4 D2"]
            summary_df["TEMPLATE 5 D1"] = df["TEMPLATE 5 D1"]
            summary_df["TEMPLATE 5 D2"] = df["TEMPLATE 5 D2"]
            summary_df["TEMPLATE 6 D1"] = df["TEMPLATE 6 D1"]
            summary_df["TEMPLATE 6 D2"] = df["TEMPLATE 6 D2"]
            if st.session_state.collector_file is not None:
                try:
                    if st.session_state.collector_file.name.endswith('.csv'):
                        collector_df = pd.read_csv(st.session_state.collector_file, encoding='utf-8-sig', skipinitialspace=True, dtype=str)
                    elif st.session_state.collector_file.name.endswith('.xlsx'):
                        collector_df = pd.read_excel(st.session_state.collector_file, engine='openpyxl', dtype=str)
                    collector_df = collector_df.fillna("")
                    with st.expander("🔍 Show Detected Column Names (Collector File)"):
                        st.write("Detected Column Names:", list(collector_df.columns))
                    collector_required_columns = ["Financing/Card No.", "Collector", "Assign Date"]
                    collector_missing_columns = [col for col in collector_required_columns if col not in collector_df.columns]
                    if collector_missing_columns:
                        st.error(f"Missing required columns in collector file: {', '.join(collector_missing_columns)}")
                        st.stop()
                    summary_df = summary_df.merge(
                        collector_df[["Financing/Card No.", "Collector", "Assign Date"]],
                        how="left",
                        left_on="Contract Number",
                        right_on="Financing/Card No."
                    )
                    summary_df["{{agentcode}}"] = summary_df["Collector"].fillna("")
                    summary_df["Assignment Date"] = summary_df["Assign Date"].fillna("")
                    summary_df = summary_df.drop(columns=["Financing/Card No.", "Collector", "Assign Date"], errors='ignore')
                    summary_df['{{agentcode}}'] = summary_df['{{agentcode}}'].apply(lambda x: 'PJHA' if x == 'SPMADRID' else x)
                    summary_df['{{ID}}'] = summary_df['{{agentcode}}'].apply(lambda x: 'BDCO' if x == 'PJHA' else 'BCCO' if x else '')
                    initial_row_count = len(summary_df)
                    summary_df = summary_df[summary_df["{{agentcode}}"].notna() & (summary_df["{{agentcode}}"] != "")]
                    if len(summary_df) < initial_row_count:
                        st.info(f"Removed {initial_row_count - len(summary_df)} rows where {{agentcode}} was blank or null.")
                    if summary_df["{{agentcode}}"].isna().any() or (summary_df["{{agentcode}}"] == "").any() or \
                       summary_df["Assignment Date"].isna().any() or (summary_df["Assignment Date"] == "").any():
                        st.warning("Some Contract Numbers did not match with Financing/Card No. in the collector file.")
                except Exception as e:
                    st.error(f"An error occurred while processing the collector file: {str(e)}")
            columns_to_check = [
                "Email", "{{chname}}", "{{OB}}", "{{MYP}}", "{{MAD}}",
                "{{OB+CF}}", "{{MAD+CF}}", "{{MYP+CF}}"
            ]
            summary_df = summary_df.dropna(subset=columns_to_check)
            summary_df = summary_df[~(summary_df[columns_to_check] == "").any(axis=1)]
            st.subheader("Summary Table")
            if not summary_df.empty:
                st.dataframe(summary_df, use_container_width=True)
                st.markdown("<br>", unsafe_allow_html=True)
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Summary"
                headers = list(summary_df.columns)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num).value = header
                for row_num, row in enumerate(summary_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        # Apply number format for amount columns with commas
                        if header in ["{{OB}}", "{{MYP}}", "{{MAD}}", "{{OB+CF}}", "{{MAD+CF}}", "{{MYP+CF}}"]:
                            try:
                                decimal_places = len(value.replace(',', '').split('.')[-1]) if '.' in value else 0
                                ws.cell(row=row_num, column=col_num).value = float(value.replace(',', '')) if value else 0.0
                                ws.cell(row=row_num, column=col_num).number_format = f'#,##0.{"0" * decimal_places}'
                            except:
                                ws.cell(row=row_num, column=col_num).value = value
                                ws.cell(row=row_num, column=col_num).number_format = '@'
                        else:
                            ws.cell(row=row_num, column=col_num).value = str(value)
                            ws.cell(row=row_num, column=col_num).number_format = '@'
                wb.save(output)
                output.seek(0)
                today = datetime.now().strftime("%B %d %Y")
                file_name = f"B2 Email blasting {today}.xlsx"
                st.download_button(
                    label="📥 Download Summary Table as Excel Workbook",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_email_summary",
                    use_container_width=True
                )
            else:
                st.warning("No rows remain after removing those with blank or None in Email, {{chname}}, {{OB}}, {{MYP}}, {{MAD}}, {{OB+CF}}, {{MAD+CF}}, or {{MYP+CF}} fields.")
        except Exception as e:
            st.error(f"An error occurred while processing the main file: {str(e)}")
    else:
        st.info("Please upload a CSV or Excel file to generate the summary table.")