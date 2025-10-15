import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def email_blast_level6_section():
    st.subheader("LEVEL 6 NEGATIVE ACCOUNTS Email Blast File Uploader")
    uploaded_file = st.file_uploader(
        "📤 Choose an Excel file",
        type=["xlsx"],
        key="level6_uploader",
        help="Upload an Excel (.xlsx) with columns: Email, Name, Product Type, Client Name, Account No., Financing/Card No."
    )
    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        st.success("File uploaded successfully!")
    if st.session_state.uploaded_file is not None:
        if st.button("🔄 Reset", help="Clear the uploaded file and reset"):
            st.session_state.uploaded_file = None
            st.rerun()
    if st.session_state.uploaded_file is not None:
        try:
            df = pd.read_excel(st.session_state.uploaded_file, engine='openpyxl', dtype=str)
            df = df.fillna("")
            # Validate Email column for '@' symbol
            initial_row_count_email = len(df)
            df = df[df['Email'].str.contains('@', na=False)]
            if len(df) < initial_row_count_email:
                st.info(f"Removed {initial_row_count_email - len(df)} rows where Email does not contain '@'.")
            required_columns = ['Email', 'Name', 'Product Type', 'Client Name', 'Account No.', 'Financing/Card No.']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error(f"Missing required columns: {', '.join(missing_columns)}")
            else:
                df = df[required_columns]
                summary_df = pd.DataFrame({
                    'Email': df['Email'],
                    '{{chname}}': df['Name'],
                    '{{product}}': df['Product Type'],
                    '{{agentcode}}': 'PJND',
                    'Client Name': df['Client Name'],
                    'Account No.': df['Account No.'],
                    'Financing/Card No.': df['Financing/Card No.']
                })
                st.write("### Processed Data")
                st.dataframe(summary_df, use_container_width=True)
                st.write("### Summary")
                st.write("Note: Values in Account No. and Financing/Card No. are preserved exactly as uploaded.")
                total_records = len(summary_df)
                unique_emails = summary_df['Email'].nunique()
                unique_names = summary_df['{{chname}}'].nunique()
                unique_products = summary_df['{{product}}'].nunique()
                unique_agents = summary_df['{{agentcode}}'].nunique()
                unique_clients = summary_df['Client Name'].nunique()
                unique_accounts = summary_df['Account No.'].nunique()
                unique_financing = summary_df['Financing/Card No.'].nunique()
                st.write(f"- **Total Records**: {total_records}")
                st.write(f"- **Unique Emails**: {unique_emails}")
                st.write(f"- **Unique Names ({{chname}})**: {unique_names}")
                st.write(f"- **Unique Products**: {unique_products}")
                st.write(f"- **Unique Agents ({{agentcode}})**: {unique_agents}")
                st.write(f"- **Unique Client Names**: {unique_clients}")
                st.write(f"- **Unique Account Numbers**: {unique_accounts}")
                st.write(f"- **Unique Financing/Card Numbers**: {unique_financing}")
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Summary"
                headers = list(summary_df.columns)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num).value = header
                for row_num, row in enumerate(summary_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num).value = str(value)
                        ws.cell(row=row_num, column=col_num).number_format = '@'
                wb.save(output)
                output.seek(0)
                today = datetime.now().strftime("%B %d %Y")
                file_name = f"Level 6 Negative Accounts Email blasting {today}.xlsx"
                st.download_button(
                    label="📥 Download Processed Excel",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_level6_summary",
                    use_container_width=True
                )
        except Exception as e:
            st.error(f"Error: {str(e)}")
    else:

        st.info("Please upload an Excel file to generate the summary table.")
